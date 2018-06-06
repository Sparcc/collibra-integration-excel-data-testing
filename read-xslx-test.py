import sys, os
sys.path.append(os.getcwd())
from CompareExcel import *
from openpyxl import load_workbook
import unittest

class CompareExcelMapping(CompareExcel):
    def buildMaps(self):
        #define general maps
        self.mapping={}
        self.mapping['schema'] = ['R','G']
        self.mapping['size'] = ['P','H']
        self.mapping['nullable'] = ['M','I']
        self.mapping['col_pos_id'] = ['F','J']
        self.mapping['frac_digs'] = ['O','K']
        self.mapping['default_value'] = ['I','L']
        self.mapping['desc'] = ['E','M']
        self.mapping['pk'] = ['N','N']
        
        #define one to many maps
        self.oneToManyMapping = []
        #1st mapping
        self.oneToManyMapping.append({})
        self.oneToManyMapping[0]['concat'] = 'A'
        self.oneToManyMapping[0]['separate'] = {}
        self.oneToManyMapping[0]['separate'] = {'schema':'B','table':'C','column':'F'}
    def verifyMapRow(self,rowNum):
        for k,v in self.mapping.items():
            d1 = self.ws["dest"][v[0]+str(rowNum)].value
            d2 = self.ws["src"][v[1]+str(rowNum)].value
            self.compareData(d1,d2,rowNum,v[0]+str(rowNum),v[1]+str(rowNum))
    def verifyOneToManyMapRow(self,rowNum):
        #process schema,table,column
        for map in self.oneToManyMapping:
            row = {} #stores in here for code readability
            row["dest"] = self.ws["dest"][map['concat']+str(rowNum)].value.split('::')
            subCol=0
            for k,v in map['separate'].items():
                row["src"] = self.ws["src"][v+str(rowNum)].value
                self.compareData(row["dest"][subCol],row["src"],rowNum,map['concat']+str(rowNum),v+str(rowNum))
                subCol+=1
    def processRow(self,rowNum):
        self.verifyOneToManyMapRow(rowNum)
        self.verifyMapRow(rowNum)

class CompareExcelBIDW(CompareExcelMapping):
    def buildMaps(self):
        #define general maps
        #define one to many maps
        self.oneToManyMapping = []
        #1st mapping
        self.oneToManyMapping.append({})
        self.oneToManyMapping[0]['concat'] = 'A'
        self.oneToManyMapping[0]['separate'] = {}
        self.oneToManyMapping[0]['separate'] = {'schema':'A','table':'C','column':'F'}
    
    def processRow(self,rowNum):
        self.verifyOneToManyMapRow(rowNum)
        
class TestCompareExcelIntegrationData(unittest.TestCase):
    def setUp(self):
        #checking base class
        self.comparer = CompareExcel()
        self.comparer.setDataLength(52)
        
        #checking oracle data
        self.oracleComparer = CompareExcelMapping()
        self.oracleComparer.setDataLength(36)
        
        #checking wherescape data
        self.bidwComparer = CompareExcelBIDW()
        self.bidwComparer.setDataLength(2597)
    def testCompareExcel(self):
        dest = "CompareExcel-test-file1.xlsx"
        src = "CompareExcel-test-file2.xlsx"
        self.comparer.setSourceFiles(dest, src)
        dest = ['A','C','D']
        src = ['A','C','F']
        self.comparer.setColumns(dest, src)
        self.comparer.compareFiles()
        self.assertEqual(True, self.comparer.testPassed)
    def testCompareExcelDataConversion(self):
        dest = "CompareExcel-test-file1.xlsx"
        src = "CompareExcel-test-file2.xlsx"
        self.comparer.setSourceFiles(dest, src)
        dest = ['A','C','D']
        src = ['A','C','F']
        self.comparer.setColumns(dest, src)
        terms = {'Y':'True','none':'False','F':'False','':'False'} #key compared to value
        for k,v in terms.items():
            self.assertEqual(self.comparer.convertToCommonTerm(k),v)
    def testOracle_Excel_calendar_table(self): 
        dest = "./oracle-data/oracle-collibra-export-calendar.xlsx" 
        src = "./oracle-data/oracle-export-calendar.xlsx" 
        self.oracleComparer.setSourceFiles(dest, src) 
        self.oracleComparer.buildMaps() 
        self.oracleComparer.compareFiles(startingRange = 2)
        self.assertEqual(True, self.oracleComparer.testPassed) 
    def testWherescape_Excel(self):
        dest = "./wherescape-collibra-export.xlsx"
        src = "./wherescape-export.xlsx"
        self.bidwComparer.setSourceFiles(dest, src)
        self.bidwComparer.buildMaps()
        self.bidwComparer.compareFiles(startingRange = 2)
    #def tearDown(self):
if __name__ == '__main__':
    unittest.main()