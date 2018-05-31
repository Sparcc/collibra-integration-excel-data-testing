from openpyxl import load_workbook
import unittest
import logging
import csv

class CompareExcel:
    destColumns = []
    srcColumns = []
    numCols = 0
    wb = {}
    ws = {}
    testPassed = True
    def convertToCommonTerm(self,v):
        v = str(v)
        returnValue = v
        for c in ('yes', 'true', 't', 'y'):
            if v.lower() == c:
                returnValue = 'True'
        for c in ('no', 'false', 'f', 'n'):
            if v.lower() == c:
                returnValue = 'False'
        for c in ("none",'none'): # there is a difference between ' and " !
            if v.lower() == c:
                returnValue = 'False'
        for c in ("",''):
            if v.lower() == c:
                returnValue = 'False'
        return returnValue
    def __init__(self, dest = "", src = "", debugging = False):
        self.debugging = debugging
        if debugging:
            if dest == "" or src == "":
                logging.warning('files not set!')
            else:
                self.setSourceFiles(dest,src)
    def setSourceFiles(self, dest, src):
        self.wb["dest"] = load_workbook(filename = dest)
        self.ws["dest"] = self.wb["dest"].active
        self.wb["src"] = load_workbook(filename = src)
        self.ws["src"] = self.wb["src"].active
    def setDataLength(self, upperRange):
        #TODO: Add code to calculate upperRange of both files and see if they are different
        self.upperRange = upperRange
    def compareFiles(self, startingRange = 2, limit='none'): #0 is null, 1 is heading
        rowNum=startingRange
        if limit=='none':
            while rowNum<self.upperRange:
                #compare dest row and src row for each col
                self.processRow(rowNum)
                rowNum+=1
        else:
            while rowNum<self.upperRange and rowNum <limit:
                #compare dest row and src row for each col
                self.processRow(rowNum)
                rowNum+=1
    def processRow(self,rowNum):
        #go through each column in row
        row = {}
        colNum = 0
        while colNum < self.numCols:
            row["dest"] = self.ws["dest"][self.destColumns[colNum]+str(rowNum)].value
            row["src"] = self.ws["src"][self.srcColumns[colNum]+str(rowNum)].value
            self.compareData(row["dest"],row["src"],rowNum,self.destColumns[colNum]+str(rowNum),self.srcColumns[colNum]+str(rowNum))
            colNum+=1
    def compareData(self,d1,d2,rowNum,d1Map,d2Map):
        d1 = self.convertToCommonTerm(d1)
        d2 = self.convertToCommonTerm(d2)
        if d1 != d2:
            self.testPassed = False
            print('MISMATCH! [{v3}|{v1}] does not match [{v4}|{v2}] on rowNum={rn}'.format(v1=d1,v2=d2,rn=rowNum,v3=d1Map,v4=d2Map))
            return False
    def setColumns(self, destColumns, srcColumns):
        for x in destColumns:
            self.destColumns.append(x)
            
        for x in srcColumns:
            self.srcColumns.append(x)
        self.numCols = len(destColumns)